# coding:utf-8
import openpyxl
import json, sys, os, chardet, codecs
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


class file_writer():
    def write_safeData(self, fileName, workbook, list_result_data):
        '''保存数据到文件,已完成测试项sheet页创建,过程数据页没有数据填充'''
        try:
            title = u'(%s)测试项' % fileName
            sheet1 = self.read_temple(workbook, title)
            for result_data in list_result_data:
                safeData = result_data.get('desc_Data')
                if safeData:
                    sheet1.cell(coordinate=result_data.get('result_index'), value=u'不通过')
                    sheet1.cell(coordinate=result_data.get('desc_index'), value=safeData)
                else:
                    sheet1.cell(coordinate=result_data.get('result_index'), value=u'通过')
        except:
            s = sys.exc_info()
            print 'Error "%s" happened on line %d' % (s[1], s[2].tb_lineno)

    def read_temple(self, workbook, sheetName):
        '''复制模版用例到新文件'''
        try:
            wb = openpyxl.load_workbook(u'预装APP项目测试报告.xlsx')
            ws = wb.get_sheet_by_name(u'(某APP名)测试项')
            newWB = workbook
            newWS = newWB.active
            newWS.title = sheetName
            for row in xrange(1, len(ws.rows)):
                for col in xrange(1, len(ws.columns) + 1):
                    col_to_letter = openpyxl.utils.get_column_letter(col)
                    newWS.column_dimensions[col_to_letter].width = ws.column_dimensions[col_to_letter].width
                    cell = newWS.cell(row=row, column=col, value=ws.cell(row=row, column=col).value)
                    cell.border = Border(left=Side(border_style='thin', color='FF000000'),
                                         right=Side(border_style='thin', color='FF000000'),
                                         top=Side(border_style='thin', color='FF000000'),
                                         bottom=Side(border_style='thin', color='FF000000'))
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    if row == 1:
                        cell.font = Font(name=u'宋体', bold=True, color='FFFFFFFF', size=11)
                        cell.fill = PatternFill(fill_type='solid', start_color='FF000000', end_color='FF000000')
                    else:
                        cell.font = Font(name=u'宋体', size=9)
                        if row % 2 == 0:
                            cell.fill = PatternFill(fill_type='solid', start_color='FFD9D9D9', end_color='FFD9D9D9')
            return newWS
        except:
            s = sys.exc_info()
            print 'Error "%s" happened on line %d' % (s[1], s[2].tb_lineno)
            return None

    def get_safeResult(self, filename):
        try:
            safeResult = ''
            with open(filename, 'rb') as f:
                tmp = f.read().replace('\n', '')
                safeResult = json.loads(tmp)
            return safeResult
        except:
            s = sys.exc_info()
            print 'Error "%s" happened on line %d' % (s[1], s[2].tb_lineno)

    def save_file(self, path, device, appInfo, wb, list_performance_result):
        try:
            app_name = appInfo['app_name']
            safe_result_name = device + 'Result.txt'
            safeResult = self.get_safeResult(os.path.join(path, safe_result_name))

            data = None
            if safeResult.has_key(app_name):
                data = safeResult[app_name]
            safe_result_data = {'result_index': 'G8',
                                'desc_index': 'H8',
                                'desc_Data': data}
            list_result = [safe_result_data]
            list_result.extend(list_performance_result)
            self.write_safeData(app_name, wb, list_result_data=list_result)

        except:
            s = sys.exc_info()
            print 'Error "%s" happened on line %d' % (s[1], s[2].tb_lineno)


if __name__ == '__main__':
    writer = file_writer()
    wb = openpyxl.Workbook()
    writer.save_file('123', 'hh22hh', appInfo={'name': u'开罗名门口袋学院2'}, wb=wb)
    wb.save('%s.xlsx' % u'开罗名门口袋学院2')
