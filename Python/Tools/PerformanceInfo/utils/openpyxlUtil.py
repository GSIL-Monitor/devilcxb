# -*- coding: utf-8 -*-
__author__ = 'wei.y'
import openpyxl
import openpyxl.chart as chart

column_dict = {'1': 'A',
               '2': 'B',
               '3': 'C',
               '4': 'D',
               '5': 'E',
               '6': 'F',
               '7': 'G',
               '8': 'H',
               '9': 'I'}


class WriteReport(object):
    def __init__(self, datas, wb, save_sheet_name=u'过程数据'):
        self.sheet = wb.create_sheet(title=save_sheet_name)
        self.datas = datas

    def write_data(self):
        print self.datas
        for i in xrange(1, len(self.datas) + 1):
            for j in xrange(1, len(self.datas[i - 1]) + 1):
                self.sheet.cell(row=j, column=i, value=self.datas[i - 1][j - 1])

    def add_Chart(self, data_col_index=2, chart_title='My Chart', x_title="Size", y_title="Date", chart_location='D5'):
        refObj = openpyxl.chart.Reference(self.sheet, min_col=data_col_index, min_row=1, max_col=data_col_index,
                                          max_row=len(self.datas[0]))
        # data = Reference(ws, min_col=2, min_row=1, max_col=4, max_row=7)
        chartObj = chart.LineChart()
        chartObj.style = 12
        chartObj.title = chart_title
        chartObj.y_axis.title = x_title
        chartObj.y_axis.crossAx = 500
        chartObj.y_axis.crossAx = 100
        chartObj.x_axistitle = y_title
        chartObj.add_data(refObj, titles_from_data=True)
        dates = openpyxl.chart.Reference(self.sheet, min_col=1, min_row=2, max_row=len(self.datas[0]))
        chartObj.set_categories(dates)
        self.sheet.add_chart(chartObj, chart_location)


if __name__ == '__main__':
    wb = openpyxl.Workbook()
    datas = [['1', '2', '3', '4', '5', '6', '7'],
             ['1', '2', '3', '4', '5', '6', '7'],
             ['1', '2', '3', '4', '5', '6', '7'],
             ['1', '2', '3', '4', '5', '6', '7']]
    WriteReport(datas, wb).write_data()
    wb.save('sampleChart.xlsx')
